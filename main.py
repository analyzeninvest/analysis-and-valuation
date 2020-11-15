#!/usr/bin/env python

def search_for_excel(ticker):
    import os, re
    path = "/home/arnashree/analyzeninvest-projects/NSE_Financial_Database/excel_path/"
    for files in os.listdir(path):
        xlsx_name = ticker + ".xlsx"
        if re.match(xlsx_name, files):
            return(path + xlsx_name)


def make_comparison_excel(comparison_name, list_of_stocks, dict_of_attribute):
    import pandas as pd
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl import load_workbook
    import string
    path = "/home/arnashree/analyzeninvest-projects/analysis-and-valuation/comparison_path/"
    excel_path = path + comparison_name + ".xlsx"
    wb = Workbook()
    wb.save(excel_path)
    ######################
    # make the info page #
    ######################
    writer = pd.ExcelWriter(excel_path, engine = 'openpyxl')
    df_info_stocks = pd.DataFrame(data={"Stocks":list_of_stocks})
    df_info_stocks.to_excel(writer, sheet_name = "info", startcol=0, index=False)
    #print(df_info_stocks)
    # name of parameters
    all_attributes = []
    for key in dict_of_attribute:
        all_attributes += dict_of_attribute[key]
    df_info_attributes = pd.DataFrame(data={"Parameters": all_attributes})
    df_info_attributes.to_excel(writer, sheet_name = "info", startcol=1, index=False)
    #print(df_info_attributes)
    ################################
    # make the page for all stocks #
    ################################
    #writer = pd.ExcelWriter(excel_path, engine = 'openpyxl')
    today = date.today()
    current_year = today.year
    array_of_stocks_with_attributes = fetch_attributes_from_excel(list_of_stocks, dict_of_attribute)
    for stock in list_of_stocks:
        print("################")
        print("Starting for " + stock)
        print("################")
        years_array = []
        uniqe_items_list = []
        for i in range(0,19):
            years_array.append(current_year - i)
        index = {"Year": years_array}
        dict_stock_attribute = index
        for items in array_of_stocks_with_attributes:
            for key in items:
                if key == stock:
                    uniqe_items_list.append(key)
                    dict_stock_attribute.update(items[key])
        #print(dict_stock_attribute)
        df_stock = pd.DataFrame(data = dict_stock_attribute)
        df_stock.set_index("Year")
        df_stock_T = df_stock.T
        df_stock_T.to_excel(writer, sheet_name = stock)
    writer.save()
    #######################
    # make the chart page #
    #######################
    # all stock names
    # cell = ws['B1']
    # cell.value = "Stock Name"
    # for i in range(0,len(list_of_stocks)):
    #     cell_target = 'B' + str(2 + i)
    #     cell = ws[cell_target]
    #     cell.value = list_of_stocks[i]
    # wb.save(excel_path)
    # years
    dict_stock_formula = {"Year":years_array}
    #print(len(years_array))
    # construct index match formula
    # sample :
    # =INDEX($BDL.B2:B3,MATCH(A2,$info.B2:B3))
    # =INDEX($BDL.C2:C3,MATCH(A2,$info.B2:B3))
    # =INDEX($ASTRAMICRO.B2:B3,MATCH(A2,$info.B2:B3))
    # =INDEX($ASTRAMICRO.C2:C3,MATCH(A2,$info.B2:B3))
    for stock in list_of_stocks:
        array_of_stock_formula = []
        for char in string.ascii_uppercase:
            if char not in ['A', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                formula = " = INDEX($" + stock + ".$" + char + "$3:$" + char + "$" + str(len(uniqe_items_list) +2) + ",MATCH(A2, $info.$B$2:$B$" + str(len(uniqe_items_list) +1) + ",0)" + ")"
                #print(formula)
                array_of_stock_formula.append(formula)
        #print(len(array_of_stock_formula))
        dict_stock_formula.update({stock:array_of_stock_formula})
    #print(dict_stock_formula)
    df_stock_formula = pd.DataFrame(data = dict_stock_formula)
    df_stock_formula.set_index("Year")
    df_stock_formula_T = df_stock_formula.T
    df_stock_formula_T.to_excel(writer, sheet_name = "Chart", startcol=2)
    writer.save()
    wb = load_workbook(excel_path)
    #print(wb)
    for ws in wb:
        if ws.title == "Chart":
            cell = ws['A1']
            # data validation for all the parameters
            cell.value = "Parameter"
            formula = '=$info.$B$2:$B$'+str(len(uniqe_items_list)+1)
            #print(formula)
            data_val = DataValidation(type="list",formula1=formula)
            ws.add_data_validation(data_val)
            data_val.add(ws["A2"])
    wb.save(excel_path)
    # make chart
    # make the chart

        

def fetch_attributes_from_excel(list_of_stocks, dict_of_attribute):
    """
    This will make the industry comparison excel from the existing excel sheets.
    """
    from openpyxl import load_workbook
    stock_sheet_attribute_details = []
    for stock in list_of_stocks:
        print("################")
        print("Fetching for " + stock)
        print("################")
        stock_xls_path = search_for_excel(stock)
        if not stock_xls_path:
            sys.error("Stock neither can be found nor downloaded. Exiting !!!")
        wb = load_workbook(stock_xls_path)
        for sheet_name in dict_of_attribute:
            # need to open the xls by sheet_name name.
            for sheet in wb:
                if sheet.title == sheet_name:
                    # choosing BZ for the last col from row 2
                    #cell_headers = sheet['A2':'BZ2']
                    #print(cell_headers)
                    #for cell in cell_headers:
                    for attribute in dict_of_attribute[sheet_name]:
                        #print(cell) download all lines from this
                        # attribute.  will have to go for the
                        # formatted xlsx ? no need as this will work
                        # for rawdata.
                        for col in sheet.iter_cols(min_row=2,max_col=65,max_row=2):
                            #print(col)
                            for cell in col:
                                #print(cell.value)
                                if cell.value == attribute:
                                    column = cell.column_letter
                                    #print(column)
                                    value = []
                                    key = sheet[column + str(2)].value
                                    for index in range(3, 22):
                                        raw_value = sheet[column + str(index)].value
                                        if raw_value == "--":
                                            raw_value = 0
                                        else:
                                            raw_value = raw_value.replace(",", "")
                                            raw_value = float(raw_value)
                                        value.append(raw_value) # modify here for value change in format
                                    key_value_pair = {key:value}
                                    stock_sheet_attribute_details.append({stock: key_value_pair})
                                    #print(key_value_pair)
    return(stock_sheet_attribute_details)
                        
def add_attribute():
    pass

def add_stocks():
    pass

def main():
    """
    For now this is only for testing the codes.
    """
    print("Running ... ")
    stock_ticker_list = ['BDL', 'ASTRAMICRO']
    dict_of_attribute = {"Standalone_Balance_Sheet":[
        "Total Non-Current Liabilities",
        "Current Investments Unquoted Book Value",
        "Equity Share Capital"],
                         "Standalone_Profit_and_Loss":[
                             "Total Revenue",
                             "Depreciation And Amortisation Expenses"
                         ]}
    #print(fetch_attributes_from_excel(stock_ticker_list, dict_of_attribute))
    make_comparison_excel("Defence",stock_ticker_list, dict_of_attribute)

if __name__ == '__main__':
    main()
